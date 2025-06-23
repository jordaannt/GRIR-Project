import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from pathlib import Path
from datetime import datetime
import os

def load_data(ekbe_path, po_doc_path, contacts_path):
    """Load data from Excel files."""
    ekbe = pd.read_excel(ekbe_path)
    po_doc = pd.read_excel(po_doc_path)
    contacts_df = pd.read_excel(contacts_path)
    contacts_df['Plant'] = contacts_df['Plant'].astype(str).str.strip()
    return ekbe, po_doc, contacts_df

def process_data(ekbe, po_doc):
    """Process and clean the data."""
    # Split EKBE data
    mseg = ekbe[ekbe['Trans./event type'] == 1].copy()
    rseg = ekbe[ekbe['Trans./event type'] == 2].copy()

    # Standardize key columns
    po_doc['Purchasing Document'] = po_doc['Purchasing Document'].astype(str)
    po_doc['Item'] = po_doc['Item'].astype(str).str.zfill(5)
    
    # Exclude deleted PO lines
    if 'Deletion indicator' in po_doc.columns:
        deleted_po_lines = po_doc[po_doc['Deletion indicator'] == 'L'][['Purchasing Document', 'Item']].drop_duplicates()
        po_doc = po_doc[po_doc['Deletion indicator'] != 'L']
    else:
        deleted_po_lines = None
    
    # Use the correct column for PO numbers - check if 'Purchasing Document' exists in EKBE
    if 'Purchasing Document' in mseg.columns:
        mseg['PO_Number'] = mseg['Purchasing Document'].astype(str)
        rseg['PO_Number'] = rseg['Purchasing Document'].astype(str)
    else:
        # If not, use Reference Document but we need to map it properly
        mseg['PO_Number'] = mseg['Reference Document'].astype(str)
        rseg['PO_Number'] = rseg['Reference Document'].astype(str)
    
    mseg['Item'] = mseg['Item'].astype(str).str.zfill(5)
    rseg['Item'] = rseg['Item'].astype(str).str.zfill(5)
    
    if deleted_po_lines is not None and not deleted_po_lines.empty:
        mseg = mseg.merge(deleted_po_lines, left_on=['PO_Number', 'Item'], right_on=['Purchasing Document', 'Item'], how='left', indicator=True)
        mseg = mseg[mseg['_merge'] == 'left_only'].drop(columns=['_merge', 'Purchasing Document'], errors='ignore')
        rseg = rseg.merge(deleted_po_lines, left_on=['PO_Number', 'Item'], right_on=['Purchasing Document', 'Item'], how='left', indicator=True)
        rseg = rseg[rseg['_merge'] == 'left_only'].drop(columns=['_merge', 'Purchasing Document'], errors='ignore')

    # Adjust for debit/credit
    rseg['Signed Quantity'] = rseg.apply(lambda x: -x['Quantity'] if x['Debit/Credit ind'] == 'H' else x['Quantity'], axis=1)
    rseg['Signed Amount'] = rseg.apply(lambda x: -x['Amount'] if x['Debit/Credit ind'] == 'H' else x['Amount'], axis=1)
    mseg['Signed Quantity'] = mseg.apply(lambda x: -x['Quantity'] if x['Debit/Credit ind'] == 'H' else x['Quantity'], axis=1)
    mseg['Signed Amount'] = mseg.apply(lambda x: -x['Amt.in loc.cur.'] if x['Debit/Credit ind'] == 'H' else x['Amt.in loc.cur.'], axis=1)
    
    return mseg, rseg, po_doc

def summarize_data(mseg, rseg):
    """Summarize Goods Receipt (GR) and Invoice Receipt (IR) data."""
    gr_summary = mseg.groupby(['PO_Number', 'Item', 'Material', 'Plant'], dropna=False).agg(
        GR_Qty=('Signed Quantity', 'sum'),
        GR_Value=('Signed Amount', 'sum')
    ).reset_index()
    
    ir_summary = rseg.groupby(['PO_Number', 'Item', 'Material'], dropna=False).agg(
        IR_Qty=('Signed Quantity', 'sum'),
        IR_Value=('Signed Amount', 'sum')
    ).reset_index()
    
    return gr_summary, ir_summary

def merge_summaries(gr_summary, ir_summary, po_doc):
    """Merge GR, IR, and PO data."""
    summary = pd.merge(gr_summary, ir_summary, on=['PO_Number', 'Item', 'Material'], how='outer')
    
    # Rename PO_Number to Purchasing Document for consistency
    summary = summary.rename(columns={'PO_Number': 'Purchasing Document'})
    summary['Purchasing Document'] = summary['Purchasing Document'].astype(str)
    summary['Item'] = summary['Item'].astype(str).str.zfill(5)
    
    summary = pd.merge(summary, po_doc[['Purchasing Document', 'Item', 'Short Text']], on=['Purchasing Document', 'Item'], how='left')
    
    summary = summary[[
        'Purchasing Document', 'Item', 'Material', 'Short Text', 'Plant',
        'GR_Qty', 'IR_Qty', 'GR_Value', 'IR_Value'
    ]].rename(columns={
        'Purchasing Document': 'PO', 'Item': 'Line', 'Material': 'Line/Shade', 'Short Text': 'Description',
        'GR_Qty': 'Goods Receipt Qty', 'IR_Qty': 'Invoice Qty', 'GR_Value': 'Goods Receipt Value', 'IR_Value': 'Invoice Receipt Value'
    }).sort_values(by=['PO', 'Line'])
    
    summary['Line'] = summary['Line'].astype(int)
    summary[['Goods Receipt Qty', 'Invoice Qty', 'Goods Receipt Value', 'Invoice Receipt Value']] = summary[['Goods Receipt Qty', 'Invoice Qty', 'Goods Receipt Value', 'Invoice Receipt Value']].fillna(0)
    summary['Plant'] = summary['Plant'].fillna("")
    
    return summary

def apply_issue_logic(summary, price_tolerance=0.05):
    """Apply logic to identify issues."""
    summary['Action'] = ""
    for po, group in summary.groupby('PO'):
        all_gr_no_ir = (group['Goods Receipt Qty'] > 0) & (group['Invoice Qty'] == 0)
        some_gr_no_ir = all_gr_no_ir.any() and not all_gr_no_ir.all()

        for idx, row in group.iterrows():
            gr_qty, ir_qty = row['Goods Receipt Qty'], row['Invoice Qty']
            gr_val, ir_val = row['Goods Receipt Value'], row['Invoice Receipt Value']

            if all_gr_no_ir.all():
                if idx == group.index[0]:
                    summary.at[idx, 'Action'] = "Invoice has not been paid. If you have received this order, send the invoice to AP. If you haven't received this order, contact Trade Ops to cancel the PO."
            elif some_gr_no_ir and gr_qty > 0 and ir_qty == 0:
                summary.at[idx, 'Action'] = "Short supply or credit note detected. Contact Trade Ops to reverse the goods receipt."
            elif gr_qty > ir_qty:
                summary.at[idx, 'Action'] = "You have goods receipted a higher quantity than you have been invoiced for. Contact Trade Ops to reverse the goods receipt and rectify. If this order has been split across multiple invoices, send all invoices to AP."
            elif gr_qty < ir_qty:
                summary.at[idx, 'Action'] = "You have goods receipted a lower quantity than you have been invoiced for. Amend the quantity in the Purchase Order tile and goods receipt again. If a credit note is required instead, contact the supplier to request a credit."
            elif abs(gr_val - ir_val) > price_tolerance:
                summary.at[idx, 'Action'] = "There is a discrepancy between the PO and the invoice price. Verify whether supplier pricing is correct and notify Trade Ops."
    return summary

def format_excel_file(file_path):
    """Format Excel file with styles and merged cells."""
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    header = [cell.value for cell in ws[1]]
    po_col = header.index("PO") + 1 if "PO" in header else 0
    action_col = header.index("Action") + 1 if "Action" in header else 0
    
    if po_col and action_col:
        last_po = None
        fill_idx = 0
        fill_colors = [
            PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
            PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ]
        
        po_groups = []
        current_po_start = 2
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=po_col).value != last_po:
                if last_po is not None:
                    po_groups.append({'start': current_po_start, 'end': row - 1})
                last_po = ws.cell(row=row, column=po_col).value
                current_po_start = row
                fill_idx = (fill_idx + 1) % len(fill_colors)
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_colors[fill_idx]
                cell.border = thin_border
                
                # Format currency and apply red text for issues
                col_header = ws.cell(row=1, column=col).value
                if col_header in ['Goods Receipt Value', 'Invoice Receipt Value']:
                    cell.number_format = '$#,##0.00'
                
                # Only make Action column text red when there's an action
                action_value = ws.cell(row=row, column=action_col).value
                if action_value and action_value.strip() and col == action_col:
                    cell.font = Font(color="FF0000")

        if last_po is not None:
            po_groups.append({'start': current_po_start, 'end': ws.max_row})
        
        # Merge action cells for "not paid" issues
        for group in po_groups:
            action_value = ws.cell(row=group['start'], column=action_col).value
            if action_value and action_value.startswith("Invoice has not been paid"):
                if group['end'] > group['start']:
                    ws.merge_cells(start_row=group['start'], start_column=action_col, end_row=group['end'], end_column=action_col)
                    ws.cell(row=group['start'], column=action_col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    
    wb.save(file_path)

def generate_email_reports(summary, contacts_df, send_emails=False):
    """Generate reports and optionally send emails."""
    issues_only = summary[summary['Action'].notna() & (summary['Action'] != "")]
    plant_po_issues = defaultdict(str)
    
    for plant, plant_group in issues_only.groupby('Plant'):
        report = f"<h2>GRIR Report {datetime.now().strftime('%B')} - {plant}</h2>"
        for po, po_group in plant_group.groupby('PO'):
            report += f"<br><b style='background-color: #FFFF00;'>PO {po}</b><br>"
            if po_group['Action'].str.startswith("Invoice has not been paid").all():
                report += f"<span style='color: red;'>{po_group['Action'].iloc[0]}</span><br><br>"
            else:
                for _, row in po_group.iterrows():
                    report += (
                        f"Line {int(row['Line'])} | {row['Line/Shade']} | {row['Description']}<br>"
                        f"You goods receipted: {int(row['Goods Receipt Qty'])} @ ${row['Goods Receipt Value']:.2f}<br>"
                        f"You were invoiced: {int(row['Invoice Qty'])} @ ${row['Invoice Receipt Value']:.2f}<br>"
                        f"<span style='color: red;'>{row['Action']}</span></b><br><br>"
                    )
        plant_po_issues[plant] = report

    if send_emails:
        # Get email settings from environment variables
        SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
        SENDER_EMAIL = os.getenv('SENDER_EMAIL', 'duluxtradeops@gmail.com')
        SENDER_PASSWORD = os.getenv('SENDER_PASSWORD', 'gljy uctw cfcn cqrz')
        
        # Check if environment variables are set
        if SENDER_PASSWORD == 'gljy uctw cfcn cqrz':
            print("Warning: Using default password. Set SENDER_PASSWORD environment variable for production use.")
        
        for _, row in contacts_df.iterrows():
            plant, to_email = row['Plant'], row['Email']
            cc_email = row['CC'] if pd.notna(row['CC']) else ""
            message_body = plant_po_issues.get(plant)

            if message_body:
                plant_df = summary[summary['Plant'] == plant]
                plant_file = Path(f"GRIR_Report_{plant}.xlsx")
                plant_df.to_excel(plant_file, index=False)
                format_excel_file(plant_file)

                html_body = f"<html><body>{message_body}<p><i>Attached is the full GRIR report for your plant.</i></p></body></html>"
                msg = MIMEMultipart()
                msg['From'], msg['To'], msg['CC'], msg['Subject'] = SENDER_EMAIL, to_email, cc_email, f"GRIR Report – {plant}"
                msg.attach(MIMEText(html_body, 'html'))

                with open(plant_file, "rb") as f:
                    part = MIMEApplication(f.read(), Name=plant_file.name)
                    part['Content-Disposition'] = f'attachment; filename="{plant_file.name}"'
                    msg.attach(part)
                
                recipients = [to_email] + ([cc_email] if cc_email else [])
                try:
                    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                        server.starttls()
                        server.login(SENDER_EMAIL, SENDER_PASSWORD)
                        server.sendmail(SENDER_EMAIL, recipients, msg.as_string())
                    print(f"Sent GRIR summary to {to_email} (cc: {cc_email})")
                except Exception as e:
                    print(f"Failed to send email to {to_email}: {e}")
            else:
                print(f"No issues for plant {plant}, no email sent.")

def run_analysis(ekbe_path, po_doc_path, contacts_path, output_summary_path="GRIR_summary.xlsx", send_emails=False, price_tolerance=0.05):
    """Main function to run the full GRIR analysis."""
    # Load data
    ekbe, po_doc, contacts_df = load_data(ekbe_path, po_doc_path, contacts_path)
    
    # Process and summarize data
    mseg, rseg, po_doc = process_data(ekbe, po_doc)
    gr_summary, ir_summary = summarize_data(mseg, rseg)
    
    # Merge and finalize summary
    summary = merge_summaries(gr_summary, ir_summary, po_doc)
    
    # Apply issue logic
    summary = apply_issue_logic(summary, price_tolerance)
    
    # Save and format main summary file
    summary.to_excel(output_summary_path, index=False)
    format_excel_file(output_summary_path)
    
    # Generate reports and emails
    generate_email_reports(summary, contacts_df, send_emails)
    
    print("GRIR analysis complete.")
    return summary

if __name__ == '__main__':
    # Define file paths
    EKBE_PATH = "data/EKBE.xlsx"
    PO_DOC_PATH = "data/EKPO.XLSX"
    CONTACTS_PATH = "data/email.xlsx"
    
    # Run the analysis
    run_analysis(EKBE_PATH, PO_DOC_PATH, CONTACTS_PATH, send_emails=True)
